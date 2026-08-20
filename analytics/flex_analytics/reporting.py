"""Excel reporting for standalone Flex analytics."""

from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Any

from .config import DEFAULT_REPORT_PATH, EXCEL_FEED_DIR
from .excel_feeds import FEED_SCHEMAS, export_excel_feeds, build_excel_feed_data


class AnalyticsReportError(RuntimeError):
    pass


def _write_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]], fields: tuple[str, ...]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(list(fields))
    for row in rows:
        sheet.append([row.get(field) for field in fields])


def generate_excel_report(conn: sqlite3.Connection, path: Path = DEFAULT_REPORT_PATH) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise AnalyticsReportError("openpyxl is not installed") from exc
    path.parent.mkdir(parents=True, exist_ok=True)
    data = build_excel_feed_data(conn)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    sheet_names = {
        "daily_nav": "NAV_Curve",
        "cumulative_return": "Cumulative_Return",
        "pnl_by_symbol": "PnL_By_Symbol",
        "commissions": "Commissions",
        "executions": "Executions",
    }
    for feed_name, rows in data.items():
        _write_sheet(workbook, sheet_names[feed_name], rows, FEED_SCHEMAS[feed_name].fields)
    workbook.save(path)
    return path


def ensure_trading_performance_workbook(path: Path = DEFAULT_REPORT_PATH) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise AnalyticsReportError("openpyxl is not installed") from exc
    if path.exists():
        return path
    workbook = Workbook()
    workbook.active.title = "TradingbotR1000"
    workbook.save(path)
    return path


def export_report_and_feeds(conn: sqlite3.Connection) -> dict[str, Any]:
    return {
        "workbook": str(generate_excel_report(conn)),
        "feeds": [str(path) for path in export_excel_feeds(conn, EXCEL_FEED_DIR)],
    }
