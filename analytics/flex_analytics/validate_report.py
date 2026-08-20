from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - depends on local runtime environment.
    load_workbook = None

from .config import DEFAULT_REPORT_PATH
from .excel_feeds import build_excel_feed_data


MONEY_TOLERANCE = 0.01
PERCENT_TOLERANCE_DECIMAL = 0.000001


class ReportValidationError(RuntimeError):
    pass


def _as_iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value or "").strip()
    if " " in text:
        text = text.split(" ", 1)[0]
    if "T" in text:
        text = text.split("T", 1)[0]
    return text


def _as_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def _rows_by_header(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise ReportValidationError("openpyxl is not installed")
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "") for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue
        rows.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return rows


def _latest(rows: list[dict[str, Any]], date_field: str) -> dict[str, Any]:
    if not rows:
        raise ReportValidationError(f"No rows for {date_field}.")
    return max(rows, key=lambda row: _as_iso_date(row.get(date_field)))


def _compare_money(label: str, expected: Any, actual: Any, failures: list[str]) -> None:
    if abs(_as_float(expected) - _as_float(actual)) > MONEY_TOLERANCE:
        failures.append(f"{label}: expected {expected}, workbook {actual}")


def _compare_percent(label: str, expected: Any, actual: Any, failures: list[str]) -> None:
    if abs(_as_float(expected) - _as_float(actual)) > PERCENT_TOLERANCE_DECIMAL:
        failures.append(f"{label}: expected {expected}, workbook {actual}")


def validate_report_accuracy(conn: sqlite3.Connection, workbook_path: Path = DEFAULT_REPORT_PATH) -> dict[str, Any]:
    feeds = build_excel_feed_data(conn)
    failures: list[str] = []

    workbook_daily_nav = _rows_by_header(workbook_path, "NAV_Curve")
    workbook_cumulative = _rows_by_header(workbook_path, "Cumulative_Return")
    workbook_symbol = _rows_by_header(workbook_path, "PnL_By_Symbol")
    workbook_commissions = _rows_by_header(workbook_path, "Commissions")
    workbook_executions = _rows_by_header(workbook_path, "Executions")

    expected_nav = _latest(feeds["daily_nav"], "report_date")
    actual_nav = _latest(workbook_daily_nav, "report_date")
    latest_statement_date = str(expected_nav["report_date"])

    _compare_money("starting NAV", expected_nav["starting_nav"], actual_nav.get("starting_nav"), failures)
    _compare_money("ending NAV", expected_nav["ending_nav"], actual_nav.get("ending_nav"), failures)

    expected_cumulative = _latest(feeds["cumulative_return"], "report_date")
    actual_cumulative = _latest(workbook_cumulative, "report_date")
    _compare_money("cumulative P&L", expected_cumulative["cumulative_pnl"], actual_cumulative.get("cumulative_pnl"), failures)
    _compare_percent("cumulative return", expected_cumulative["cumulative_return_pct"], actual_cumulative.get("cumulative_return_pct"), failures)

    expected_symbols = {
        (str(row["report_date"]), str(row["symbol"]), str(row["conid"])): row
        for row in feeds["pnl_by_symbol"]
        if str(row["report_date"]) == latest_statement_date
    }
    actual_symbols = {
        (_as_iso_date(row.get("report_date")), str(row.get("symbol") or ""), str(row.get("conid") or "")): row
        for row in workbook_symbol
        if _as_iso_date(row.get("report_date")) == latest_statement_date
    }
    if set(expected_symbols) != set(actual_symbols):
        failures.append(f"P&L by symbol keys differ: expected {sorted(expected_symbols)}, workbook {sorted(actual_symbols)}")
    for key, expected in expected_symbols.items():
        actual = actual_symbols.get(key)
        if not actual:
            continue
        _compare_money(f"P&L by symbol net {key}", expected["net_pnl_base"], actual.get("net_pnl_base"), failures)
        _compare_money(f"P&L by symbol commissions {key}", expected["commissions_base"], actual.get("commissions_base"), failures)
        _compare_percent(f"P&L by symbol pct {key}", expected["pnl_pct"], actual.get("pnl_pct"), failures)
        if "UNKNOWN" in key:
            failures.append(f"Unexplained UNKNOWN P&L row remains: {key}")

    expected_commissions = sum(_as_float(row["commission_base"]) for row in feeds["commissions"])
    actual_commissions = sum(_as_float(row.get("commission_base")) for row in workbook_commissions)
    _compare_money("commissions total", expected_commissions, actual_commissions, failures)

    expected_execution_count = len(feeds["executions"])
    actual_execution_count = len(workbook_executions)
    if expected_execution_count != actual_execution_count:
        failures.append(f"execution count: expected {expected_execution_count}, workbook {actual_execution_count}")

    expected_latest_execution = max((str(row["trade_date"]) for row in feeds["executions"]), default="")
    actual_latest_execution = max((_as_iso_date(row.get("trade_date")) for row in workbook_executions), default="")
    if expected_latest_execution != actual_latest_execution:
        failures.append(f"latest execution date: expected {expected_latest_execution}, workbook {actual_latest_execution}")

    if failures:
        raise ReportValidationError("; ".join(failures))

    return {
        "latest_statement_date": latest_statement_date,
        "latest_execution_date": expected_latest_execution,
        "execution_count": expected_execution_count,
        "commissions_total": expected_commissions,
        "status": "PASSED",
    }
