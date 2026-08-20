# manual_control_console.py

import asyncio
import csv
from contextlib import nullcontext
from datetime import datetime, timezone
import json
import math
import os
from pathlib import Path
import socket
import threading
import time

try:
    asyncio.get_running_loop()
except RuntimeError:
    asyncio.set_event_loop(asyncio.new_event_loop())

from ib_insync import IB, ExecutionFilter, LimitOrder, MarketOrder, Stock

import config as cfg
try:
    from execution_history import (
        format_execution_history_lines,
        format_latest_execution_history_lines,
        load_latest_execution_history,
        load_execution_history_since_baseline,
    )
except ImportError:
    from execution_history import load_latest_execution_history as _load_r1000_execution_rows

    def load_latest_execution_history(limit: int = 20, **_kwargs):
        rows = _load_r1000_execution_rows(limit=limit)
        return rows, {
            "source": "r1000_execution_history",
            "total_rows": len(rows),
            "coverage_status": "AVAILABLE" if rows else "EMPTY",
        }

    def load_execution_history_since_baseline(limit: int = 20, **_kwargs):
        return load_latest_execution_history(limit=limit)

    def _simple_execution_line(row: dict) -> str:
        return " | ".join(
            [
                str(row.get("timestamp_utc") or row.get("time") or "-"),
                str(row.get("symbol") or "-"),
                str(row.get("side") or "-"),
                str(row.get("quantity") or "-"),
                str(row.get("price") or "-"),
                str(row.get("reason") or "-"),
                str(row.get("source") or "-"),
            ]
        )

    def format_latest_execution_history_lines(rows: list[dict], metadata: dict) -> list[str]:
        lines = [
            "Source: R1000 execution-history store",
            f"Rows shown: {len(rows)} of {metadata.get('total_rows', len(rows))} stored executions",
        ]
        if not rows:
            lines.append("No stored broker executions are available.")
            return lines
        lines.append("Time | Symbol | Side | Qty | Price | Reason | Source")
        lines.extend(_simple_execution_line(row) for row in rows)
        return lines

    def format_execution_history_lines(rows: list[dict], metadata: dict) -> list[str]:
        return format_latest_execution_history_lines(rows, metadata)
from ibkr_utils import (
    IBKR_BLOCKING_REQUEST_TIMEOUT_SECONDS,
    MARKET_HOURS_TIME_UNAVAILABLE_REASON,
    get_contract_details,
    get_ibkr_server_time,
    get_zoneinfo,
    ibkr_request_timeout,
    parse_ibkr_hours_segment,
)
from investable_capital_control import (
    InvestableCapitalControlError,
    evaluate as evaluate_investable_capital_control,
    format_usd,
    set_auto as set_investable_capital_auto,
    set_manual as set_manual_investable_capital,
)
from order_safety import LongOnlyOrderRejected, acquire_order_intent_guard


MANUAL_CLIENT_ID = int(
    os.environ.get(
        "TRADINGBOTR1000_MANUAL_CLIENT_ID",
        os.environ.get("TRADINGBOT_MANUAL_CLIENT_ID", cfg.MANUAL_CLIENT_ID),
    )
)
MANUAL_CONNECT_ATTEMPTS = int(
    os.environ.get(
        "TRADINGBOTR1000_MANUAL_CONNECT_ATTEMPTS",
        os.environ.get("TRADINGBOT_MANUAL_CONNECT_ATTEMPTS", "3"),
    )
)
MANUAL_CONNECT_RETRY_DELAY_SECONDS = float(
    os.environ.get(
        "TRADINGBOTR1000_MANUAL_CONNECT_RETRY_DELAY_SECONDS",
        os.environ.get("TRADINGBOT_MANUAL_CONNECT_RETRY_DELAY_SECONDS", "2"),
    )
)
EXECUTIONS_REQUEST_TIMEOUT_SECONDS = float(
    os.environ.get(
        "TRADINGBOTR1000_EXECUTIONS_REQUEST_TIMEOUT_SECONDS",
        os.environ.get(
            "TRADINGBOT_EXECUTIONS_REQUEST_TIMEOUT_SECONDS",
            str(IBKR_BLOCKING_REQUEST_TIMEOUT_SECONDS),
        ),
    )
)
MANUAL_ACTIONS_LOG = Path(cfg.BASE_DIR) / "logs" / "manual_actions.log"
MANUAL_EXPORTS_DIR = Path(cfg.BASE_DIR) / "reports" / "manual_console"
STOP_BOT_FILE = cfg.STOP_FILE
PROJECT_DIR = cfg.PROJECT_ROOT
PROJECT_CONFIG_DIR = PROJECT_DIR / "config"
MANUAL_WATCHLIST_XLSX = PROJECT_CONFIG_DIR / "manual_trading_watchlist.xlsx"
MANUAL_WATCHLIST_CSV = PROJECT_CONFIG_DIR / "manual_trading_watchlist.csv"

US_PRIMARY_EXCHANGES = {
    "NASDAQ",
    "NYSE",
    "AMEX",
    "ARCA",
    "BATS",
    "IEX",
    "NYSEARCA",
    "NYSEMKT",
}

DEFAULT_MANUAL_WATCHLIST = [
    ("AAPL", "Apple Inc.", "NASDAQ", "USD", ""),
    ("ABBV", "AbbVie Inc.", "NYSE", "USD", ""),
    ("ABT", "Abbott Laboratories", "NYSE", "USD", ""),
    ("AMD", "Advanced Micro Devices, Inc.", "NASDAQ", "USD", ""),
    ("AMZN", "Amazon.com, Inc.", "NASDAQ", "USD", ""),
    ("AVGO", "Broadcom Inc.", "NASDAQ", "USD", ""),
    ("BAC", "Bank of America Corporation", "NYSE", "USD", ""),
    ("BRK B", "Berkshire Hathaway Inc. Class B", "NYSE", "USD", ""),
    ("COST", "Costco Wholesale Corporation", "NASDAQ", "USD", ""),
    ("CRM", "Salesforce, Inc.", "NYSE", "USD", ""),
    ("DIS", "The Walt Disney Company", "NYSE", "USD", ""),
    ("GOOG", "Alphabet Inc. Class C", "NASDAQ", "USD", ""),
    ("GOOGL", "Alphabet Inc. Class A", "NASDAQ", "USD", ""),
    ("HD", "The Home Depot, Inc.", "NYSE", "USD", ""),
    ("IBM", "International Business Machines Corporation", "NYSE", "USD", ""),
    ("INTC", "Intel Corporation", "NASDAQ", "USD", ""),
    ("JNJ", "Johnson & Johnson", "NYSE", "USD", ""),
    ("JPM", "JPMorgan Chase & Co.", "NYSE", "USD", ""),
    ("KO", "The Coca-Cola Company", "NYSE", "USD", ""),
    ("LLY", "Eli Lilly and Company", "NYSE", "USD", ""),
    ("MA", "Mastercard Incorporated", "NYSE", "USD", ""),
    ("META", "Meta Platforms, Inc.", "NASDAQ", "USD", ""),
    ("MRK", "Merck & Co., Inc.", "NYSE", "USD", ""),
    ("MSFT", "Microsoft Corporation", "NASDAQ", "USD", ""),
    ("NFLX", "Netflix, Inc.", "NASDAQ", "USD", ""),
    ("NVDA", "NVIDIA Corporation", "NASDAQ", "USD", ""),
    ("ORCL", "Oracle Corporation", "NYSE", "USD", ""),
    ("PEP", "PepsiCo, Inc.", "NASDAQ", "USD", ""),
    ("PG", "The Procter & Gamble Company", "NYSE", "USD", ""),
    ("QQQ", "Invesco QQQ Trust", "NASDAQ", "USD", ""),
    ("SPY", "SPDR S&P 500 ETF Trust", "ARCA", "USD", ""),
    ("TSLA", "Tesla, Inc.", "NASDAQ", "USD", ""),
    ("UNH", "UnitedHealth Group Incorporated", "NYSE", "USD", ""),
    ("V", "Visa Inc.", "NYSE", "USD", ""),
    ("WMT", "Walmart Inc.", "NYSE", "USD", ""),
    ("XOM", "Exxon Mobil Corporation", "NYSE", "USD", ""),
]

_manual_log_lock = threading.RLock()


class ManualControlError(RuntimeError):
    pass


class BrokerActionError(ManualControlError):
    pass


class OutsideLiquidHoursError(ManualControlError):
    pass


FINAL_ORDER_STATES = {
    "Filled",
    "Cancelled",
    "ApiCancelled",
    "Inactive",
}
SUBMITTED_ORDER_STATES = {
    "ApiPending",
    "PendingSubmit",
    "PreSubmitted",
    "Submitted",
}
CANCELLED_ORDER_STATES = {
    "Cancelled",
    "ApiCancelled",
    "PendingCancel",
}
ERROR_ORDER_STATES = {
    "Inactive",
}


def _utc_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace(
        "+00:00", "Z"
    )


def _plain_number(value):
    try:
        number = float(value)
        return number if math.isfinite(number) else None
    except Exception:
        return None


def _order_identifiers(trade) -> dict:
    order = getattr(trade, "order", None)
    status = getattr(trade, "orderStatus", None)
    contract = getattr(trade, "contract", None)
    return {
        "order_id": getattr(order, "orderId", None),
        "perm_id": getattr(order, "permId", None),
        "client_id": getattr(order, "clientId", None),
        "ibkr_status": getattr(status, "status", None),
        "filled": _plain_number(getattr(status, "filled", None)),
        "remaining": _plain_number(getattr(status, "remaining", None)),
        **contract_identity(contract),
    }


def contract_identity(contract) -> dict:
    return {
        "symbol": getattr(contract, "symbol", None),
        "con_id": getattr(contract, "conId", None),
        "currency": getattr(contract, "currency", None),
        "exchange": getattr(contract, "exchange", None),
        "primary_exchange": getattr(contract, "primaryExchange", None),
    }


def contract_identity_text(contract_or_details) -> str:
    details = (
        contract_identity(contract_or_details)
        if not isinstance(contract_or_details, dict)
        else contract_or_details
    )
    return (
        f"{details.get('symbol') or '-'} | "
        f"conId={details.get('con_id') or '-'} | "
        f"currency={details.get('currency') or '-'} | "
        f"exchange={details.get('exchange') or '-'} | "
        f"primary={details.get('primary_exchange') or '-'}"
    )


def classify_trade_outcome(trade, timed_out: bool = False) -> dict:
    status = getattr(getattr(trade, "orderStatus", None), "status", None) or "UNKNOWN"
    log_entries = list(getattr(trade, "log", None) or [])
    error_messages = [
        str(getattr(entry, "message", ""))
        for entry in log_entries
        if str(getattr(entry, "message", "")).strip()
    ]
    has_error_log = any(
        keyword in message.lower()
        for message in error_messages
        for keyword in ("error", "reject", "rejected", "cancelled", "inactive")
    )

    if status == "Filled":
        category = "Filled"
        success = True
        source = "broker filled"
    elif status in CANCELLED_ORDER_STATES:
        category = "Cancelled"
        success = False
        source = "cancelled/rejected by IBKR or user"
    elif status in ERROR_ORDER_STATES or has_error_log:
        category = "Rejected/Error"
        success = False
        source = "rejected by IBKR"
    elif status in SUBMITTED_ORDER_STATES:
        category = "Submitted"
        success = True
        source = "submitted but inactive/held"
    elif timed_out:
        category = "Unknown/Timed out"
        success = False
        source = "unknown/timed out"
    else:
        category = "Unknown/Timed out"
        success = False
        source = "unknown/timed out"

    return {
        "category": category,
        "success": success,
        "outcome_source": source,
        "ibkr_status": status,
        "timed_out": bool(timed_out),
        "error_messages": error_messages,
        **_order_identifiers(trade),
    }


def log_manual_action(event: str, status: str = "OK", **details) -> bool:
    """Append one structured audit record for console or future remote callers."""
    record = {
        "timestamp": _utc_timestamp(),
        "action": str(event),
        "status": str(status),
        "details": details,
    }
    line = json.dumps(record, default=str, sort_keys=True)

    try:
        with _manual_log_lock:
            MANUAL_ACTIONS_LOG.parent.mkdir(parents=True, exist_ok=True)
            with open(MANUAL_ACTIONS_LOG, "a", encoding="utf-8") as handle:
                handle.write(line + "\n")
                handle.flush()
                os.fsync(handle.fileno())
        return True
    except Exception as error:
        print(
            "CRITICAL: manual action audit log unavailable | "
            f"error={type(error).__name__}"
        )
        return False


def _require_audit(event: str, **details) -> None:
    if not log_manual_action(event, "REQUESTED", **details):
        raise ManualControlError(
            "Action refused because manual_actions.log could not be written."
        )


def _post_broker_audit(event: str, status: str, **details) -> bool:
    try:
        return bool(log_manual_action(event, status, **details))
    except Exception as error:
        print(
            "BROKER ACTION COMPLETED BUT AUDIT/LOCAL LOGGING FAILED | "
            f"event={event} | error={type(error).__name__}"
        )
        return False


def _combine_details(*items: dict) -> dict:
    combined = {}
    for item in items:
        combined.update(item)
    return combined


def _manual_connection_details() -> dict:
    return {
        "host": cfg.HOST,
        "port": cfg.PORT,
        "client_id": MANUAL_CLIENT_ID,
        "attempts": MANUAL_CONNECT_ATTEMPTS,
        "retry_delay_seconds": MANUAL_CONNECT_RETRY_DELAY_SECONDS,
    }


def _ibkr_socket_probe(timeout_seconds: float = 2.0) -> tuple[bool, str]:
    try:
        with socket.create_connection((cfg.HOST, cfg.PORT), timeout=timeout_seconds):
            return True, "reachable"
    except Exception as error:
        return False, type(error).__name__


def connect_manual_console(ib: IB) -> None:
    if ib.isConnected():
        return

    connection_details = _manual_connection_details()
    _require_audit("CONNECT", **connection_details)

    last_error = None
    attempts = max(1, MANUAL_CONNECT_ATTEMPTS)
    for attempt in range(1, attempts + 1):
        socket_ok, socket_status = _ibkr_socket_probe()
        log_manual_action(
            "CONNECT_ATTEMPT",
            "REQUESTED",
            **connection_details,
            attempt=attempt,
            socket_reachable=socket_ok,
            socket_status=socket_status,
        )
        try:
            ib.connect(
                cfg.HOST,
                cfg.PORT,
                clientId=MANUAL_CLIENT_ID,
                timeout=10,
            )
        except Exception as error:
            last_error = error
            log_manual_action(
                "CONNECT_ATTEMPT",
                "FAILED",
                **connection_details,
                attempt=attempt,
                error_type=type(error).__name__,
            )
            if attempt < attempts:
                time.sleep(max(0.0, MANUAL_CONNECT_RETRY_DELAY_SECONDS))
            continue

        if ib.isConnected():
            log_manual_action("CONNECT", "COMPLETED", client_id=MANUAL_CLIENT_ID)
            return

        last_error = ManualControlError("IBKR connection did not become active.")
        log_manual_action(
            "CONNECT_ATTEMPT",
            "FAILED",
            **connection_details,
            attempt=attempt,
            error_type="NotConnected",
        )
        if attempt < attempts:
            time.sleep(max(0.0, MANUAL_CONNECT_RETRY_DELAY_SECONDS))

    error_type = type(last_error).__name__ if last_error is not None else "Unknown"
    log_manual_action("CONNECT", "FAILED", **connection_details, error_type=error_type)
    raise ManualControlError(
        "Could not connect to IBKR "
        f"on {cfg.HOST}:{cfg.PORT} with manual client ID {MANUAL_CLIENT_ID} "
        f"after {attempts} attempt(s): {error_type}. "
        "Verify IB Gateway Paper API is enabled/listening and that the manual client ID "
        f"{MANUAL_CLIENT_ID} is not already in use."
    ) from last_error


def _preferred_account_value(summary, tag: str):
    matches = [item for item in summary if getattr(item, "tag", None) == tag]
    if not matches:
        return {"value": None, "currency": None}
    item = next(
        (entry for entry in matches if getattr(entry, "currency", None) == "USD"),
        matches[0],
    )
    value = _plain_number(getattr(item, "value", None))
    return {"value": value, "currency": getattr(item, "currency", None)}


def get_account_summary(ib: IB) -> dict:
    summary = ib.accountSummary()
    result = {
        "net_liquidation": _preferred_account_value(summary, "NetLiquidation"),
        "cash": _preferred_account_value(summary, "TotalCashValue"),
        "buying_power": _preferred_account_value(summary, "BuyingPower"),
        "unrealized_pnl": _preferred_account_value(summary, "UnrealizedPnL"),
    }
    log_manual_action("VIEW_ACCOUNT_SUMMARY", fields=list(result))
    return result


def _default_watchlist_rows() -> list[dict]:
    return [
        {
            "ticker": ticker,
            "company_name": company_name,
            "exchange": exchange,
            "currency": currency,
            "notes": notes,
        }
        for ticker, company_name, exchange, currency, notes in DEFAULT_MANUAL_WATCHLIST
    ]


def _normalize_watchlist_row(row: dict) -> dict | None:
    ticker = str(row.get("ticker", "") or "").strip().upper()
    if not ticker:
        return None
    return {
        "ticker": ticker,
        "company_name": str(row.get("company_name", "") or "").strip(),
        "exchange": str(row.get("exchange", "") or "").strip().upper(),
        "currency": str(row.get("currency", "") or "USD").strip().upper(),
        "notes": str(row.get("notes", "") or "").strip(),
    }


def _openpyxl_module():
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except Exception:
        return None


def create_manual_watchlist_template() -> Path:
    PROJECT_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    rows = _default_watchlist_rows()
    openpyxl = _openpyxl_module()
    if openpyxl is not None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "manual_watchlist"
        fields = ["ticker", "company_name", "exchange", "currency", "notes"]
        sheet.append(fields)
        for row in rows:
            sheet.append([row.get(field, "") for field in fields])
        workbook.save(MANUAL_WATCHLIST_XLSX)
        return MANUAL_WATCHLIST_XLSX

    with MANUAL_WATCHLIST_CSV.open("w", newline="", encoding="utf-8") as handle:
        fields = ["ticker", "company_name", "exchange", "currency", "notes"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return MANUAL_WATCHLIST_CSV


def load_manual_watchlist() -> tuple[list[dict], Path, list[str]]:
    warnings = []
    if not MANUAL_WATCHLIST_XLSX.exists() and not MANUAL_WATCHLIST_CSV.exists():
        created = create_manual_watchlist_template()
        warnings.append(f"Manual trading watchlist template created: {created}")

    rows = []
    source = MANUAL_WATCHLIST_XLSX if MANUAL_WATCHLIST_XLSX.exists() else MANUAL_WATCHLIST_CSV
    if source.suffix.lower() == ".xlsx":
        openpyxl = _openpyxl_module()
        if openpyxl is None:
            warnings.append(
                "openpyxl is not installed; Excel watchlist cannot be read. "
                "Using CSV fallback template instead."
            )
            if not MANUAL_WATCHLIST_CSV.exists():
                create_manual_watchlist_template()
            source = MANUAL_WATCHLIST_CSV
        else:
            try:
                workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
                sheet = workbook.active
                header = [
                    str(cell or "").strip().lower()
                    for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                ]
                for values in sheet.iter_rows(min_row=2, values_only=True):
                    raw = {
                        header[index]: values[index]
                        for index in range(min(len(header), len(values)))
                    }
                    normalized = _normalize_watchlist_row(raw)
                    if normalized is not None:
                        rows.append(normalized)
            except Exception as error:
                warnings.append(
                    f"Excel watchlist could not be read ({type(error).__name__}); "
                    "using built-in defaults for this session."
                )
                rows = _default_watchlist_rows()
    if source.suffix.lower() == ".csv":
        try:
            with source.open("r", newline="", encoding="utf-8-sig") as handle:
                for raw in csv.DictReader(handle):
                    normalized = _normalize_watchlist_row(raw)
                    if normalized is not None:
                        rows.append(normalized)
        except Exception as error:
            warnings.append(
                f"CSV watchlist could not be read ({type(error).__name__}); "
                "using built-in defaults for this session."
            )
            rows = _default_watchlist_rows()

    if not rows:
        warnings.append("Manual trading watchlist is empty; using built-in defaults.")
        rows = _default_watchlist_rows()

    deduped = {}
    for row in rows:
        deduped[row["ticker"]] = row
    return sorted(deduped.values(), key=lambda item: item["ticker"]), source, warnings


def print_manual_watchlist(rows: list[dict], source: Path, warnings: list[str]) -> None:
    print(f"\nMANUAL TRADING WATCHLIST ({source})")
    for warning in warnings:
        print(f"WARNING: {warning}")
    print("#   TICKER    COMPANY NAME                         PRIMARY      CUR  NOTES")
    for index, row in enumerate(rows, 1):
        print(
            f"{index:<3} {row['ticker']:<9} "
            f"{row['company_name'][:35]:<35} "
            f"{(row.get('exchange') or '-'):<12} "
            f"{(row.get('currency') or 'USD'):<4} "
            f"{row.get('notes') or ''}"
        )


def get_positions(ib: IB) -> list[dict]:
    positions = []
    for item in ib.portfolio():
        contract = item.contract
        quantity = _plain_number(getattr(item, "position", None))
        if quantity is None or quantity == 0:
            continue
        positions.append(
            {
                "contract": contract,
                **contract_identity(contract),
                "quantity": quantity,
                "average_cost": _plain_number(getattr(item, "averageCost", None)),
                "market_price": _plain_number(getattr(item, "marketPrice", None)),
                "market_value": _plain_number(getattr(item, "marketValue", None)),
                "unrealized_pnl": _plain_number(getattr(item, "unrealizedPNL", None)),
            }
        )
    log_manual_action("VIEW_POSITIONS", count=len(positions))
    return positions


def get_open_trades(ib: IB):
    request_all = getattr(ib, "reqAllOpenOrders", None)
    raw_trades = list(request_all() if callable(request_all) else ib.openTrades())
    trades = _unique_active_open_trades(raw_trades)
    log_manual_action(
        "VIEW_OPEN_ORDERS",
        count=len(trades),
        raw_count=len(raw_trades),
    )
    return trades


def _open_order_key(trade):
    order = getattr(trade, "order", None)
    contract = getattr(trade, "contract", None)
    perm_id = getattr(order, "permId", None)
    if perm_id not in (None, "", 0):
        return ("perm_id", str(perm_id))
    order_id = getattr(order, "orderId", None)
    if order_id not in (None, "", 0):
        return ("order_id", str(order_id))
    return (
        "fallback",
        getattr(contract, "conId", None),
        getattr(contract, "symbol", None),
        getattr(order, "action", None),
        getattr(order, "orderType", None),
        getattr(order, "totalQuantity", None),
        getattr(order, "lmtPrice", None),
    )


def _is_terminal_or_not_open(summary: dict) -> bool:
    status = str(summary.get("status") or "")
    remaining = summary.get("remaining")
    return (
        status in FINAL_ORDER_STATES
        or status in CANCELLED_ORDER_STATES
        or remaining == 0
    )


def _unique_active_open_trades(trades):
    grouped: dict[tuple, list] = {}
    for trade in trades:
        grouped.setdefault(_open_order_key(trade), []).append(trade)

    active = []
    for group in grouped.values():
        summaries = [summarize_open_trade(trade) for trade in group]
        if any(_is_terminal_or_not_open(summary) for summary in summaries):
            continue
        active.append(group[-1])
    return active


def _contract_matches(left, right) -> bool:
    left_con_id = getattr(left, "conId", None)
    right_con_id = getattr(right, "conId", None)
    if left_con_id and right_con_id:
        return left_con_id == right_con_id
    return (
        getattr(left, "symbol", "") == getattr(right, "symbol", "")
        and getattr(left, "secType", "") == getattr(right, "secType", "")
        and getattr(left, "currency", "") == getattr(right, "currency", "")
    )


def refresh_broker_state(ib: IB) -> None:
    """Ask IBKR for fresh account/order state without failing the console flow."""
    for method_name in ("reqPositions", "reqAllOpenOrders"):
        method = getattr(ib, method_name, None)
        if callable(method):
            try:
                method()
            except Exception:
                pass
    sleep = getattr(ib, "sleep", None)
    if callable(sleep):
        sleep(0.5)


def _load_manual_order_sources() -> dict:
    sources = {}
    if not MANUAL_ACTIONS_LOG.exists():
        return sources
    try:
        for line in MANUAL_ACTIONS_LOG.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            try:
                record = json.loads(line)
            except Exception:
                continue
            if record.get("action") != "PLACE_ORDER":
                continue
            details = record.get("details") or {}
            order_type = details.get("order_type")
            for key_name in ("order_id", "perm_id"):
                key_value = details.get(key_name)
                if key_value is not None:
                    sources[(key_name, str(key_value))] = {
                        "source": "manual console",
                        "order_type": order_type,
                    }
    except Exception:
        return sources
    return sources


def _execution_source(execution, manual_sources: dict) -> tuple[str, str | None]:
    order_id = getattr(execution, "orderId", None)
    perm_id = getattr(execution, "permId", None)
    for key in (("order_id", str(order_id)), ("perm_id", str(perm_id))):
        if key in manual_sources:
            item = manual_sources[key]
            return item.get("source", "manual console"), item.get("order_type")
    client_id = getattr(execution, "clientId", None)
    if client_id in {MANUAL_CLIENT_ID, getattr(cfg, "MANUAL_CLIENT_ID", None)}:
        return "manual console", None
    if client_id == getattr(cfg, "CLIENT_ID", None):
        return "bot", None
    if client_id in {
        getattr(cfg, "RECONCILIATION_CLIENT_ID", None),
        getattr(cfg, "REMOTE_CONTROL_CLIENT_ID", None),
    }:
        return "unknown", None
    if client_id is not None:
        return "external", None
    return "unknown", None


def _current_day_execution_filter():
    exec_filter = ExecutionFilter()
    try:
        exec_filter.clientId = 0
    except Exception:
        pass
    return exec_filter


def _request_current_day_executions(
    ib: IB,
    timeout_seconds: float = EXECUTIONS_REQUEST_TIMEOUT_SECONDS,
) -> list:
    exec_filter = _current_day_execution_filter()
    try:
        with ibkr_request_timeout(ib, timeout_seconds):
            return list(ib.reqExecutions(exec_filter))
    except (TimeoutError, asyncio.TimeoutError) as error:
        raise ManualControlError(
            "Timed out waiting for IBKR execDetailsEnd() after "
            f"{timeout_seconds:g} seconds."
        ) from error


def _execution_dedupe_key(row: dict) -> tuple:
    exec_id = str(row.get("execution_id") or "").strip()
    if exec_id:
        return ("execId", exec_id)
    return (
        "composite",
        str(row.get("timestamp") or ""),
        str(row.get("symbol") or ""),
        str(row.get("side") or ""),
        str(row.get("quantity") or ""),
        str(row.get("average_execution_price") or ""),
        str(row.get("order_id") or ""),
        str(row.get("perm_id") or ""),
        str(row.get("client_id") or ""),
    )


def _execution_sort_key(row: dict) -> str:
    return str(row.get("timestamp") or "")


def get_recent_executions(ib: IB, limit: int = 20) -> list[dict]:
    limit = max(1, int(limit))
    fills = _request_current_day_executions(ib)
    manual_sources = _load_manual_order_sources()
    rows_by_key = {}
    for fill in fills:
        contract = getattr(fill, "contract", None)
        execution = getattr(fill, "execution", None)
        if execution is None:
            continue
        source, order_type = _execution_source(execution, manual_sources)
        side_raw = str(getattr(execution, "side", "") or "").upper()
        side = "BUY" if side_raw in {"BOT", "BUY"} else "SELL" if side_raw in {
            "SLD",
            "SELL",
        } else side_raw or "UNKNOWN"
        row = {
            "timestamp": str(getattr(execution, "time", "") or ""),
            **contract_identity(contract),
            "side": side,
            "quantity": _plain_number(getattr(execution, "shares", None)),
            "average_execution_price": _plain_number(
                getattr(execution, "avgPrice", None)
            )
            or _plain_number(getattr(execution, "price", None)),
            "order_type": order_type,
            "source": source,
            "order_id": getattr(execution, "orderId", None),
            "perm_id": getattr(execution, "permId", None),
            "execution_id": getattr(execution, "execId", None),
            "client_id": getattr(execution, "clientId", None),
        }
        rows_by_key.setdefault(_execution_dedupe_key(row), row)
    rows = list(rows_by_key.values())
    rows.sort(key=_execution_sort_key, reverse=True)
    log_manual_action("VIEW_RECENT_EXECUTIONS", count=min(limit, len(rows)))
    return rows[:limit]


def summarize_open_trade(trade) -> dict:
    order = trade.order
    status = trade.orderStatus
    return {
        "trade": trade,
        "order_id": getattr(order, "orderId", None),
        "perm_id": getattr(order, "permId", None),
        **contract_identity(trade.contract),
        "action": getattr(order, "action", ""),
        "order_type": getattr(order, "orderType", ""),
        "quantity": _plain_number(getattr(order, "totalQuantity", None)),
        "limit_price": _plain_number(getattr(order, "lmtPrice", None)),
        "status": getattr(status, "status", ""),
        "filled": _plain_number(getattr(status, "filled", None)),
        "remaining": _plain_number(getattr(status, "remaining", None)),
    }


def _is_us_listed_usd_stock(contract) -> bool:
    currency = str(getattr(contract, "currency", "") or "").upper()
    sec_type = str(getattr(contract, "secType", "") or "").upper()
    primary = str(getattr(contract, "primaryExchange", "") or "").upper()
    exchange = str(getattr(contract, "exchange", "") or "").upper()
    return (
        currency == "USD"
        and sec_type == "STK"
        and (primary in US_PRIMARY_EXCHANGES or exchange in US_PRIMARY_EXCHANGES)
    )


def _contract_company_name(contract_details) -> str:
    for attr in ("longName", "marketName"):
        value = getattr(contract_details, attr, None)
        if value:
            return str(value)
    return ""


def _contract_candidates_for_symbol(ib: IB, symbol: str) -> list[dict]:
    request = getattr(ib, "reqContractDetails", None)
    if callable(request):
        details = request(Stock(symbol, "SMART", "USD"))
        candidates = []
        for item in details:
            contract = getattr(item, "contract", None)
            if contract is None or not _is_us_listed_usd_stock(contract):
                continue
            candidates.append(
                {
                    "contract": contract,
                    "company_name": _contract_company_name(item),
                }
            )
        return candidates

    qualified = ib.qualifyContracts(Stock(symbol, "SMART", "USD"))
    return [
        {"contract": contract, "company_name": ""}
        for contract in qualified
        if _is_us_listed_usd_stock(contract)
    ]


def print_contract_candidates(candidates: list[dict]) -> None:
    print("\nAVAILABLE USD / US-LISTED CONTRACTS")
    print("#  SYMBOL     CONID      CUR  EXCHANGE     PRIMARY      COMPANY")
    for index, item in enumerate(candidates, 1):
        contract = item["contract"]
        identity = contract_identity(contract)
        print(
            f"{index:<2} {identity.get('symbol') or '-':<10} "
            f"{str(identity.get('con_id') or '-'):<9} "
            f"{identity.get('currency') or '-':<4} "
            f"{identity.get('exchange') or '-':<12} "
            f"{identity.get('primary_exchange') or '-':<12} "
            f"{item.get('company_name') or ''}"
        )


def resolve_us_listed_usd_contract(ib: IB, symbol: str, input_fn=input):
    symbol = symbol.strip().upper()
    if not symbol:
        raise ManualControlError("Symbol is required.")

    candidates = _contract_candidates_for_symbol(ib, symbol)
    if not candidates:
        raise ManualControlError(
            f"No valid USD / US-listed stock or ETF contract found for {symbol}. "
            "Release 2 Manual Control Console trading is restricted to USD "
            "US-listed stocks/ETFs."
        )

    unique = {}
    for item in candidates:
        contract = item["contract"]
        key = getattr(contract, "conId", None) or (
            getattr(contract, "symbol", None),
            getattr(contract, "currency", None),
            getattr(contract, "primaryExchange", None),
        )
        unique[key] = item
    candidates = list(unique.values())

    if len(candidates) == 1:
        return qualify_manual_contract(ib, candidates[0]["contract"])

    print_contract_candidates(candidates)
    try:
        selection = int(input_fn("Contract number to use: ").strip())
    except Exception as error:
        raise ManualControlError("Explicit contract selection is required.") from error
    if selection < 1 or selection > len(candidates):
        raise ManualControlError("Contract selection is out of range.")
    return qualify_manual_contract(ib, candidates[selection - 1]["contract"])


def qualify_stock(ib: IB, symbol: str, currency: str = "USD"):
    currency = currency.strip().upper()
    if currency != "USD":
        raise ManualControlError(
            "Release 2 Manual Control Console trading is restricted to USD "
            "US-listed stocks/ETFs. Non-USD manual BUY/SELL is refused."
        )
    return resolve_us_listed_usd_contract(ib, symbol)


def qualify_manual_contract(ib: IB, contract):
    """Qualify supported manual-trading contracts with SMART routing where possible."""
    symbol = getattr(contract, "symbol", "").strip().upper()
    currency = (getattr(contract, "currency", "") or "USD").strip().upper()
    sec_type = (getattr(contract, "secType", "") or "STK").strip().upper()
    if sec_type not in {"STK", "ETF"}:
        raise ManualControlError(
            f"Unsupported manual contract type {sec_type or 'UNKNOWN'}; "
            "Release 2 supports US-listed stocks/ETFs only."
        )
    if not symbol:
        raise ManualControlError("Contract symbol is required.")
    if getattr(contract, "conId", None):
        qualified = ib.qualifyContracts(contract)
        if qualified:
            return qualified[0]
    smart = Stock(symbol, "SMART", currency)
    primary = getattr(contract, "primaryExchange", None)
    if primary:
        smart.primaryExchange = primary
    qualified = ib.qualifyContracts(smart)
    if not qualified:
        raise ManualControlError(f"IBKR could not qualify {symbol} {currency}.")
    return qualified[0]


def _hours_open(hours_text, timezone_id, now=None):
    if not hours_text or not timezone_id or now is None:
        return None
    tz = get_zoneinfo(timezone_id)
    if tz is None:
        return None
    current = now.replace(tzinfo=timezone.utc) if now.tzinfo is None else now.astimezone(timezone.utc)
    now_local = current.astimezone(tz)

    parsed_any = False
    for block in hours_text.split(";"):
        for segment in block.split(","):
            parsed = parse_ibkr_hours_segment(segment.strip())
            if parsed is None:
                continue
            parsed_any = True
            start_naive, end_naive = parsed
            start = start_naive.replace(tzinfo=tz)
            end = end_naive.replace(tzinfo=tz)
            if start <= now_local <= end:
                return True
    return False if parsed_any else None


def get_market_hours_status(
    ib: IB,
    contract,
    now=None,
    *,
    time_source: str | None = None,
    allow_server_time_lookup: bool = True,
) -> dict:
    details = get_contract_details(ib, contract)
    if details is None:
        return {
            "known": False,
            "trading_open": None,
            "liquid_open": None,
            "timezone": None,
            "time_source": "UNAVAILABLE",
            "current_time": "",
            "detail": "Market-hours status unavailable: IBKR contractDetails unavailable",
        }

    timezone_id = getattr(details, "timeZoneId", None)
    current_time = now
    time_source = time_source or ("SUPPLIED_TIME" if current_time is not None else "UNAVAILABLE")
    if current_time is None and allow_server_time_lookup:
        current_time = get_ibkr_server_time(ib)
        time_source = "IBKR_SERVER_TIME" if current_time is not None else "UNAVAILABLE"
    if current_time is None:
        return {
            "known": False,
            "trading_open": None,
            "liquid_open": None,
            "timezone": timezone_id,
            "time_source": time_source,
            "current_time": "",
            "detail": MARKET_HOURS_TIME_UNAVAILABLE_REASON,
        }
    current_time = current_time.replace(tzinfo=timezone.utc) if current_time.tzinfo is None else current_time.astimezone(timezone.utc)
    trading_open = _hours_open(
        getattr(details, "tradingHours", None), timezone_id, current_time
    )
    liquid_open = _hours_open(
        getattr(details, "liquidHours", None), timezone_id, current_time
    )
    tz = get_zoneinfo(timezone_id)
    current_display = current_time.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S %Z") if tz is not None else ""
    return {
        "known": trading_open is not None or liquid_open is not None,
        "trading_open": trading_open,
        "liquid_open": liquid_open,
        "timezone": timezone_id,
        "time_source": time_source,
        "current_time": current_display,
        "detail": "",
    }


def get_current_market_price(ib: IB, contract) -> dict:
    ticker = None
    try:
        ticker = ib.reqMktData(contract, "", False, False)
        ib.sleep(1.0)
        values = {
            "market": _plain_number(
                ticker.marketPrice() if callable(getattr(ticker, "marketPrice", None)) else None
            ),
            "last": _plain_number(getattr(ticker, "last", None)),
            "close": _plain_number(getattr(ticker, "close", None)),
            "bid": _plain_number(getattr(ticker, "bid", None)),
            "ask": _plain_number(getattr(ticker, "ask", None)),
        }
        for key in ("market", "last", "ask", "bid", "close"):
            value = values.get(key)
            if value is not None and value > 0:
                return {"price": value, "source": key, **values}
        return {"price": None, "source": "unavailable", **values}
    except Exception as error:
        return {
            "price": None,
            "source": f"error:{type(error).__name__}",
            "market": None,
            "last": None,
            "close": None,
            "bid": None,
            "ask": None,
        }
    finally:
        if ticker is not None:
            try:
                ib.cancelMktData(contract)
            except Exception:
                pass


def suggest_buy_limit_price(price_info: dict) -> float | None:
    price = _plain_number(price_info.get("ask")) or _plain_number(price_info.get("price"))
    if price is None or price <= 0:
        return None
    return round(price, 2)


def _position_quantity(ib: IB, contract) -> float:
    con_id = getattr(contract, "conId", None)
    symbol = getattr(contract, "symbol", "")
    currency = getattr(contract, "currency", "")
    sec_type = getattr(contract, "secType", "")
    total = 0.0
    for item in ib.portfolio():
        item_contract = item.contract
        same_contract = con_id and getattr(item_contract, "conId", None) == con_id
        same_symbol = (
            not con_id
            and getattr(item_contract, "symbol", "") == symbol
            and getattr(item_contract, "currency", "") == currency
            and getattr(item_contract, "secType", "") == sec_type
        )
        if same_contract or same_symbol:
            total += float(item.position)
    return total


def current_broker_position(ib: IB, contract, *, refresh: bool = True) -> float:
    """Return the current broker-reported net position for a contract."""
    contract = qualify_manual_contract(ib, contract)
    if refresh:
        refresh_broker_state(ib)
    return _position_quantity(ib, contract)


def _open_sell_remaining(ib: IB, contract) -> float:
    remaining = 0.0
    for trade in get_open_trades(ib):
        trade_contract = getattr(trade, "contract", None)
        if not _contract_matches(trade_contract, contract):
            continue
        order = getattr(trade, "order", None)
        status = getattr(trade, "orderStatus", None)
        if str(getattr(order, "action", "")).upper() != "SELL":
            continue
        if getattr(status, "status", None) in FINAL_ORDER_STATES:
            continue
        status_remaining = _plain_number(getattr(status, "remaining", None))
        if status_remaining is not None:
            remaining += max(0.0, status_remaining)
            continue
        total = _plain_number(getattr(order, "totalQuantity", None)) or 0.0
        filled = _plain_number(getattr(status, "filled", None)) or 0.0
        remaining += max(0.0, total - filled)
    return remaining


def validate_long_only_sell(
    ib: IB,
    contract,
    quantity: float,
    *,
    context: str = "SELL",
    refresh: bool = True,
) -> float:
    """Refuse any manual SELL/liquidation that could create or increase a short."""
    contract = qualify_manual_contract(ib, contract)
    quantity = float(quantity)
    if not math.isfinite(quantity) or quantity <= 0:
        raise ManualControlError("Quantity must be positive.")
    current_position = current_broker_position(ib, contract, refresh=refresh)
    if current_position <= 0:
        raise ManualControlError(
            f"{context} refused for {contract_identity_text(contract)}: current "
            f"broker position is {current_position:g}. This console is long-only "
            "and will not create or increase a short position."
        )
    open_sell_remaining = _open_sell_remaining(ib, contract)
    available_long = max(0.0, current_position - open_sell_remaining)
    if quantity > available_long:
        raise ManualControlError(
            f"{context} refused for {contract_identity_text(contract)}: requested "
            f"SELL quantity {quantity:g} exceeds available long quantity "
            f"{available_long:g} after current open SELL orders "
            f"({open_sell_remaining:g}). Short selling is disabled."
        )
    return current_position


def wait_for_trade_outcome(
    ib: IB,
    trade,
    timeout_seconds: float = 8.0,
    submitted_settle_seconds: float = 1.5,
) -> dict:
    started = datetime.now(timezone.utc).timestamp()
    deadline = started + max(0.0, timeout_seconds)
    submitted_seen_at = None
    timed_out = False
    while True:
        status = getattr(getattr(trade, "orderStatus", None), "status", None)
        now = datetime.now(timezone.utc).timestamp()
        if status in FINAL_ORDER_STATES:
            break
        if status in SUBMITTED_ORDER_STATES:
            if submitted_seen_at is None:
                submitted_seen_at = now
            if now - submitted_seen_at >= submitted_settle_seconds:
                break
        if now >= deadline:
            timed_out = True
            break
        ib.sleep(0.25)
    return classify_trade_outcome(trade, timed_out=timed_out)


def place_manual_order(
    ib: IB,
    contract,
    action: str,
    quantity: float,
    order_type: str,
    limit_price: float | None = None,
    acknowledge_outside_liquid_hours: bool = False,
    allow_short: bool = False,
    sell_context: str = "SELL",
):
    contract = qualify_manual_contract(ib, contract)
    action = action.strip().upper()
    order_type = order_type.strip().upper()
    quantity = float(quantity)
    if action not in {"BUY", "SELL"}:
        raise ManualControlError("Action must be BUY or SELL.")
    if order_type not in {"LIMIT", "MARKET"}:
        raise ManualControlError("Order type must be LIMIT or MARKET.")
    if not math.isfinite(quantity) or quantity <= 0:
        raise ManualControlError("Quantity must be positive.")
    if order_type == "LIMIT":
        limit_price = float(limit_price)
        if not math.isfinite(limit_price) or limit_price <= 0:
            raise ManualControlError("Limit price must be positive.")

    current_position = None
    if action == "SELL" and not allow_short:
        current_position = validate_long_only_sell(
            ib,
            contract,
            quantity,
            context=sell_context,
            refresh=True,
        )

    hours = get_market_hours_status(ib, contract)
    if hours["liquid_open"] is not True and not acknowledge_outside_liquid_hours:
        raise OutsideLiquidHoursError(
            "Liquid-hours status is closed or unknown; explicit acknowledgement "
            "is required."
        )

    audit_details = {
        **contract_identity(contract),
        "action": action,
        "quantity": quantity,
        "order_type": order_type,
        "limit_price": limit_price if order_type == "LIMIT" else None,
        "liquid_open": hours["liquid_open"],
        "pre_order_position": current_position,
    }
    _require_audit("PLACE_ORDER", **audit_details)

    order = (
        LimitOrder(action, quantity, limit_price)
        if order_type == "LIMIT"
        else MarketOrder(action, quantity)
    )
    order.tif = "DAY"
    order.outsideRth = False

    try:
        if action == "SELL" and not allow_short:
            try:
                order_guard_context = acquire_order_intent_guard(
                    ib,
                    contract,
                    action,
                    quantity,
                    allow_short=False,
                    context=sell_context,
                    strategy="ManualControlConsole",
                    refresh=True,
                )
            except LongOnlyOrderRejected as error:
                _post_broker_audit(
                    "PLACE_ORDER",
                    "REJECTED",
                    **audit_details,
                    error_type="LongOnlyOrderRejected",
                    error=str(error),
                )
                raise ManualControlError(str(error)) from error
        else:
            order_guard_context = nullcontext(None)

        with order_guard_context as order_guard:
            trade = ib.placeOrder(contract, order)
            outcome = wait_for_trade_outcome(ib, trade)
            if order_guard is not None:
                order_guard.mark_submitted(trade)
    except Exception as error:
        _post_broker_audit(
            "PLACE_ORDER",
            "FAILED",
            **audit_details,
            error_type=type(error).__name__,
        )
        raise BrokerActionError(
            f"IBKR order placement failed: {type(error).__name__}"
        ) from error

    audit_ok = _post_broker_audit(
        "PLACE_ORDER",
        outcome["category"],
        **_combine_details(audit_details, outcome),
    )
    outcome["audit_ok"] = audit_ok
    return trade, outcome


def cancel_selected_order(ib: IB, trade):
    order = trade.order
    order_id = getattr(order, "orderId", None)
    identity = contract_identity(trade.contract)
    _require_audit("CANCEL_ORDER", order_id=order_id, **identity)
    try:
        result = ib.cancelOrder(order)
        outcome = wait_for_trade_outcome(ib, trade, timeout_seconds=5.0)
    except Exception as error:
        _post_broker_audit(
            "CANCEL_ORDER",
            "FAILED",
            order_id=order_id,
            **identity,
            error_type=type(error).__name__,
        )
        raise ManualControlError(
            f"Order cancellation failed: {type(error).__name__}"
        ) from error
    audit_ok = _post_broker_audit(
        "CANCEL_ORDER",
        outcome["category"],
        **_combine_details({"order_id": order_id, **identity}, outcome),
    )
    outcome["audit_ok"] = audit_ok
    return result, outcome


def cancel_all_orders(ib: IB):
    """Request IBKR global cancellation, including orders from other API clients."""
    _require_audit("CANCEL_ALL_ORDERS", scope="GLOBAL")
    try:
        result = ib.reqGlobalCancel()
        ib.sleep(1.0)
        remaining = [
            summarize_open_trade(trade)
            for trade in get_open_trades(ib)
            if summarize_open_trade(trade)["status"] not in CANCELLED_ORDER_STATES
        ]
    except Exception as error:
        _post_broker_audit(
            "CANCEL_ALL_ORDERS", "FAILED", error_type=type(error).__name__
        )
        raise ManualControlError(
            f"Global cancellation failed: {type(error).__name__}"
        ) from error
    status = "Completed" if not remaining else "Unknown/Timed out"
    audit_ok = _post_broker_audit(
        "CANCEL_ALL_ORDERS",
        status,
        scope="GLOBAL",
        remaining_open=len(remaining),
    )
    return result, {
        "category": status,
        "remaining_open": len(remaining),
        "audit_ok": audit_ok,
    }


def liquidate_position(
    ib: IB,
    position: dict,
    acknowledge_outside_liquid_hours: bool = False,
):
    stale_quantity = float(position["quantity"])
    symbol = position["symbol"]
    contract = qualify_manual_contract(ib, position["contract"])
    current_quantity = current_broker_position(ib, contract, refresh=True)
    if current_quantity <= 0:
        raise ManualControlError(
            f"Liquidation refused for {contract_identity_text(contract)}: current broker position is "
            f"{current_quantity:g}. The Manual Control Console is long-only and "
            "will not buy-to-cover or create/increase a short position."
        )
    open_sell_remaining = _open_sell_remaining(ib, contract)
    if open_sell_remaining > 0:
        raise ManualControlError(
            f"Liquidation refused for {contract_identity_text(contract)}: there is already an open SELL "
            f"order for {open_sell_remaining:g} shares. Refresh executions and "
            "positions before submitting another liquidation."
        )
    quantity = min(abs(stale_quantity), current_quantity)
    if quantity <= 0:
        raise ManualControlError(f"Liquidation refused for {symbol}: no long quantity.")
    _require_audit(
        "LIQUIDATE_POSITION",
        **contract_identity(contract),
        action="SELL",
        selected_quantity=stale_quantity,
        current_position=current_quantity,
        quantity=quantity,
    )
    trade, outcome = place_manual_order(
        ib,
        contract,
        "SELL",
        quantity,
        "MARKET",
        acknowledge_outside_liquid_hours=acknowledge_outside_liquid_hours,
        allow_short=False,
        sell_context="Liquidation",
    )
    try:
        outcome["post_order_position"] = current_broker_position(
            ib,
            contract,
            refresh=True,
        )
    except Exception:
        outcome["post_order_position"] = None
    return trade, outcome


def emergency_liquidate_all(
    ib: IB,
    positions: list[dict] | None = None,
    acknowledge_outside_liquid_hours: bool = False,
) -> list[dict]:
    positions = list(positions if positions is not None else get_positions(ib))
    positive_positions = []
    skipped_positions = []
    for position in positions:
        quantity = _plain_number(position.get("quantity"))
        if quantity is not None and quantity > 0:
            positive_positions.append(position)
        else:
            skipped_positions.append(position)

    _require_audit(
        "EMERGENCY_LIQUIDATE_ALL",
        position_count=len(positions),
        positive_long_count=len(positive_positions),
        skipped_non_long_count=len(skipped_positions),
    )

    try:
        STOP_BOT_FILE.write_text(
            "STOP requested by TradingbotR1000 Control Console emergency liquidation",
            encoding="utf-8",
        )
    except Exception as error:
        _post_broker_audit(
            "EMERGENCY_LIQUIDATE_ALL",
            "FAILED",
            stage="STOP_BOT",
            error_type=type(error).__name__,
        )
        raise ManualControlError(
            "Emergency liquidation refused because STOP_BOT.txt could not be created."
        ) from error

    _, cancel_outcome = cancel_all_orders(ib)
    results = []
    for position in skipped_positions:
        results.append(
            {
                "symbol": position["symbol"],
                **{
                    key: value
                    for key, value in position.items()
                    if key
                    in {
                        "con_id",
                        "currency",
                        "exchange",
                        "primary_exchange",
                    }
                },
                "trade": None,
                "ok": False,
                "error": "RefusedNonLongPosition",
            }
        )

    for position in positive_positions:
        try:
            trade, outcome = liquidate_position(
                ib,
                position,
                acknowledge_outside_liquid_hours=acknowledge_outside_liquid_hours,
            )
            results.append(
                {
                    "symbol": position["symbol"],
                    **{
                        key: value
                        for key, value in position.items()
                        if key
                        in {
                            "con_id",
                            "currency",
                            "exchange",
                            "primary_exchange",
                        }
                    },
                    "trade": trade,
                    "ok": bool(outcome["success"]),
                    "outcome": outcome,
                }
            )
        except Exception as error:
            results.append(
                {
                    "symbol": position["symbol"],
                    **{
                        key: value
                        for key, value in position.items()
                        if key
                        in {
                            "con_id",
                            "currency",
                            "exchange",
                            "primary_exchange",
                        }
                    },
                    "trade": None,
                    "ok": False,
                    "error": type(error).__name__,
                }
            )

    _post_broker_audit(
        "EMERGENCY_LIQUIDATE_ALL",
        "REQUESTS_SENT",
        cancel_outcome=cancel_outcome,
        successful=sum(1 for item in results if item["ok"]),
        failed=sum(1 for item in results if not item["ok"]),
    )
    return results


def confirm_yes_no(prompt: str, input_fn=input) -> bool:
    entered = input_fn(f"{prompt} [y/N]: ").strip().lower()
    return entered == "y"


def parse_quantity_input(value: str, *, allow_all: bool = False, all_quantity=None):
    text = value.strip().lower()
    if text == "all":
        if not allow_all:
            raise ManualControlError("Quantity 'all' is not valid for this action.")
        quantity = _plain_number(all_quantity)
        if quantity is None or quantity <= 0:
            raise ManualControlError("No positive quantity is available for 'all'.")
        return float(quantity), True
    try:
        quantity = float(text)
    except Exception as error:
        raise ManualControlError("A numeric quantity or valid 'all' is required.") from error
    if not math.isfinite(quantity) or quantity <= 0:
        raise ManualControlError("Quantity must be positive.")
    return quantity, False


def _format_value(item) -> str:
    value = item.get("value")
    currency = item.get("currency") or ""
    return "N/A" if value is None else f"{value:,.2f} {currency}".strip()


def print_account_summary(ib: IB) -> None:
    summary = get_account_summary(ib)
    print("\nACCOUNT SUMMARY")
    print(f"Net Liquidation : {_format_value(summary['net_liquidation'])}")
    print(f"Cash            : {_format_value(summary['cash'])}")
    print(f"Buying Power    : {_format_value(summary['buying_power'])}")
    print(f"Unrealized P&L  : {_format_value(summary['unrealized_pnl'])}")


def run_investable_capital_control(ib: IB) -> None:
    summary = get_account_summary(ib)
    live_nlv = summary["net_liquidation"]["value"]
    if live_nlv is None:
        raise ManualControlError("Live IBKR NetLiquidation is unavailable.")

    current = evaluate_investable_capital_control(live_nlv)
    print("\nINVESTABLE CAPITAL CONTROL")
    print(f"NLV                         : {format_usd(current['live_net_liquidation'])}")
    print(f"Investable Capital Mode     : {current['mode']}")
    print(f"Configured Investable Capital: {format_usd(current['configured_investable_capital'])}")
    print(f"Effective Investable Capital : {format_usd(current['effective_investable_capital'])}")
    print(f"Compliance                  : {current['compliance']}")
    if current.get("reason"):
        print(f"Reason                      : {current['reason']}")
    print(f"Settings file               : {current['settings_file']}")

    entered = input("Enter USD amount, AUTO, or blank to leave unchanged: ").strip()
    if not entered:
        log_manual_action("INVESTABLE_CAPITAL_CONTROL", "UNCHANGED", mode=current["mode"])
        print("Investable capital setting unchanged.")
        return
    if entered.upper() == "AUTO":
        updated = evaluate_investable_capital_control(
            live_nlv,
            settings=set_investable_capital_auto(),
        )
        log_manual_action("INVESTABLE_CAPITAL_CONTROL", "AUTO", settings_file=updated["settings_file"])
        print("Investable capital mode set to AUTO.")
    else:
        try:
            settings = set_manual_investable_capital(entered, live_net_liquidation=live_nlv)
        except InvestableCapitalControlError as error:
            log_manual_action("INVESTABLE_CAPITAL_CONTROL", "REFUSED", reason=str(error))
            raise ManualControlError(f"Invalid investable-capital amount: {error}") from error
        updated = evaluate_investable_capital_control(live_nlv, settings=settings)
        log_manual_action(
            "INVESTABLE_CAPITAL_CONTROL",
            "MANUAL",
            configured_investable_capital=updated["configured_investable_capital"],
            settings_file=updated["settings_file"],
        )
        print("Investable capital mode set to MANUAL.")

    print(f"NLV                         : {format_usd(updated['live_net_liquidation'])}")
    print(f"Investable Capital Mode     : {updated['mode']}")
    print(f"Configured Investable Capital: {format_usd(updated['configured_investable_capital'])}")
    print(f"Effective Investable Capital : {format_usd(updated['effective_investable_capital'])}")
    print(f"Compliance                  : {updated['compliance']}")


def print_positions(positions: list[dict]) -> None:
    print("\nPOSITIONS")
    if not positions:
        print("No positions.")
        return
    print(
        "#  SYMBOL       CONID      CUR  EXCHANGE     PRIMARY      "
        "QTY       AVG COST    MARKET VALUE    UNREALIZED P&L"
    )
    for index, item in enumerate(positions, 1):
        print(
            f"{index:<2} {item['symbol']:<10} {str(item.get('con_id') or '-'):<9} "
            f"{(item.get('currency') or '-'):<4} {(item.get('exchange') or '-'):<12} "
            f"{(item.get('primary_exchange') or '-'):<12} {item['quantity']:>9g} "
            f"{(item['average_cost'] or 0):>12,.2f} "
            f"{(item['market_value'] or 0):>15,.2f} "
            f"{(item['unrealized_pnl'] or 0):>17,.2f}"
        )


def print_open_orders(trades) -> None:
    print("\nOPEN ORDERS")
    if not trades:
        print("No open orders.")
        return
    print(
        "#  ID       SYMBOL     CONID      CUR  EXCHANGE     PRIMARY      "
        "ACTION TYPE      QTY       LIMIT       STATUS"
    )
    for index, trade in enumerate(trades, 1):
        item = summarize_open_trade(trade)
        limit_text = "-" if item["limit_price"] is None else f"{item['limit_price']:.4f}"
        print(
            f"{index:<2} {str(item['order_id']):<8} {item['symbol']:<10} "
            f"{str(item.get('con_id') or '-'):<9} {(item.get('currency') or '-'):<4} "
            f"{(item.get('exchange') or '-'):<12} "
            f"{(item.get('primary_exchange') or '-'):<12} "
            f"{item['action']:<6} {item['order_type']:<9} "
            f"{(item['quantity'] or 0):>8g} {limit_text:>11} {item['status']}"
        )


def print_order_outcome(prefix: str, trade, outcome: dict) -> None:
    identity = contract_identity(getattr(trade, "contract", None))
    print(
        f"{prefix}: {outcome['category']} | "
        f"Source={outcome.get('outcome_source', 'unknown')} | "
        f"{contract_identity_text(identity)} | "
        f"OrderID={outcome.get('order_id')} "
        f"PermID={outcome.get('perm_id')} "
        f"IBKRStatus={outcome.get('ibkr_status')} "
        f"Filled={outcome.get('filled')} "
        f"Remaining={outcome.get('remaining')}"
    )
    if outcome.get("error_messages"):
        print("Broker messages:")
        for message in outcome["error_messages"]:
            print(f"- {message}")
    if outcome.get("audit_ok") is False:
        print("BROKER ACTION COMPLETED BUT AUDIT/LOCAL LOGGING FAILED")


def print_recent_executions(rows: list[dict]) -> None:
    print("\nCurrent IBKR Session Executions")
    for line in format_recent_execution_lines(rows):
        print(line)


def format_recent_execution_lines(rows: list[dict]) -> list[str]:
    scope = [
        "Source: live IBKR reqExecutions()",
        "Scope: current-day executions visible to this account/API session only; use /executions for prior days.",
    ]
    if not rows:
        return [
            *scope,
            "No executions are currently available through the live IBKR API session.",
        ]
    lines = [
        *scope,
        "#  TIME                       SYMBOL CONID      CUR  EXCHANGE     "
        "PRIMARY      SIDE   QTY       AVG PRICE   TYPE     SOURCE           "
        "ORDER ID   PERM ID       EXEC ID"
    ]
    for index, row in enumerate(rows, 1):
        lines.append(
            f"{index:<2} {row['timestamp']:<26} {row['symbol']:<6} "
            f"{str(row.get('con_id') or '-'):<9} {(row.get('currency') or '-'):<4} "
            f"{(row.get('exchange') or '-'):<12} "
            f"{(row.get('primary_exchange') or '-'):<12} "
            f"{row['side']:<5} {(row['quantity'] or 0):>8g} "
            f"{(row['average_execution_price'] or 0):>11,.4f} "
            f"{(row['order_type'] or '-'):>8} {row['source']:<16} "
            f"{str(row['order_id']):<10} {str(row['perm_id']):<13} "
            f"{row['execution_id']}"
        )
    return lines


def load_current_session_executions(limit: int = 20) -> list[dict]:
    ib = IB()
    try:
        connect_manual_console(ib)
        return get_recent_executions(ib, limit)
    finally:
        try:
            if ib.isConnected():
                ib.disconnect()
        except Exception:
            pass


def print_execution_history_since_baseline(rows: list[dict], metadata: dict) -> None:
    print("Execution History Since Baseline")
    for line in format_execution_history_lines(rows, metadata):
        print(line)


def print_latest_execution_history(rows: list[dict], metadata: dict) -> None:
    print("Latest Broker Execution History")
    for line in format_latest_execution_history_lines(rows, metadata):
        print(line)


def export_recent_executions_csv(rows: list[dict]) -> Path:
    MANUAL_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = MANUAL_EXPORTS_DIR / (
        "recent_executions_" + _utc_timestamp().replace(":", "").replace("-", "") + ".csv"
    )
    fields = [
        "timestamp",
        "symbol",
        "con_id",
        "currency",
        "exchange",
        "primary_exchange",
        "side",
        "quantity",
        "average_execution_price",
        "order_type",
        "source",
        "order_id",
        "perm_id",
        "execution_id",
        "client_id",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field) for field in fields})
    log_manual_action("EXPORT_RECENT_EXECUTIONS", path=str(path), count=len(rows))
    return path


def print_market_status(status: dict, symbol: str, contract=None) -> None:
    def label(value):
        return "OPEN" if value is True else "CLOSED" if value is False else "UNKNOWN"

    print(f"\n{symbol} market-hours status")
    if contract is not None:
        print(f"Contract      : {contract_identity_text(contract)}")
    print("Source        : IBKR contractDetails tradingHours/liquidHours")
    print(f"Time source   : {status.get('time_source') or 'UNKNOWN'}")
    if status.get("current_time"):
        print(f"Current time  : {status.get('current_time')}")
    print(f"Trading hours : {label(status['trading_open'])}")
    print(f"Liquid hours  : {label(status['liquid_open'])}")
    print(f"Timezone      : {status['timezone'] or 'UNKNOWN'}")
    if status.get("detail"):
        print(f"Detail        : {status.get('detail')}")


def _market_hours_position_summary(positions: list[dict]) -> None:
    print("\nCURRENT POSITIONS FOR MARKET/LIQUID-HOURS STATUS")
    if not positions:
        print("No current owned positions.")
        return
    print(
        "#  SYMBOL     CONID      CUR  EXCHANGE     PRIMARY      "
        "QTY       AVG COST    MARKET VALUE    UNREALIZED P&L"
    )
    for index, item in enumerate(positions, 1):
        print(
            f"{index:<2} {item['symbol']:<10} {str(item.get('con_id') or '-'):<9} "
            f"{(item.get('currency') or '-'):<4} {(item.get('exchange') or '-'):<12} "
            f"{(item.get('primary_exchange') or '-'):<12} {item['quantity']:>9g} "
            f"{(item['average_cost'] or 0):>12,.2f} "
            f"{(item['market_value'] or 0):>15,.2f} "
            f"{(item['unrealized_pnl'] or 0):>17,.2f}"
        )


def _position_for_contract(contract) -> dict:
    return {
        "contract": contract,
        **contract_identity(contract),
        "quantity": 0,
        "average_cost": None,
        "market_value": None,
        "unrealized_pnl": None,
    }


def _select_market_hours_targets(ib: IB, input_fn=None) -> tuple[str, list[dict]]:
    input_fn = input_fn or input
    positions = get_positions(ib)
    _market_hours_position_summary(positions)
    choice = input_fn(
        "Select position number, type ticker, or press Enter for ALL [ALL]: "
    ).strip()

    if not choice:
        return "ALL", positions

    if choice.isdigit():
        index = int(choice)
        if index < 1 or index > len(positions):
            raise ManualControlError("Position selection is out of range.")
        return "POSITION", [positions[index - 1]]

    contract = qualify_stock(ib, choice.upper())
    return "TICKER", [_position_for_contract(contract)]


def show_market_hours_targets(ib: IB, targets: list[dict], *, continue_on_error: bool) -> dict:
    shown = 0
    failed = 0
    if not targets:
        print("No positions available for Market/Liquid-Hours Status.")
        return {"shown": shown, "failed": failed}

    server_time = get_ibkr_server_time(ib)
    time_source = "IBKR_SERVER_TIME" if server_time is not None else "UNAVAILABLE"

    for target in targets:
        contract = target.get("contract")
        symbol = target.get("symbol") or getattr(contract, "symbol", "UNKNOWN")
        try:
            status = get_market_hours_status(
                ib,
                contract,
                now=server_time,
                time_source=time_source,
                allow_server_time_lookup=False,
            )
            print_market_status(status, symbol, contract)
            log_manual_action("VIEW_MARKET_HOURS", **contract_identity(contract))
            shown += 1
        except Exception as error:
            failed += 1
            identity = contract_identity_text(contract) if contract is not None else str(symbol)
            print(
                f"\n{symbol} market-hours status ERROR: "
                f"{type(error).__name__}: {error} | {identity}"
            )
            audit_details = contract_identity(contract) if contract is not None else {}
            audit_details.update(
                {
                    "symbol": symbol,
                    "error_type": type(error).__name__,
                    "error": str(error),
                }
            )
            log_manual_action(
                "VIEW_MARKET_HOURS",
                "FAILED",
                **audit_details,
            )
            if not continue_on_error:
                raise
    return {"shown": shown, "failed": failed}


def _read_positive_number(prompt: str) -> float:
    try:
        value = float(input(prompt).strip())
    except Exception as error:
        raise ManualControlError("A numeric value is required.") from error
    if not math.isfinite(value) or value <= 0:
        raise ManualControlError("Value must be positive.")
    return value


def _read_quantity(prompt: str, *, allow_all: bool = False, all_quantity=None):
    return parse_quantity_input(
        input(prompt).strip(),
        allow_all=allow_all,
        all_quantity=all_quantity,
    )


def _select_buy_symbol_from_watchlist() -> str:
    rows, source, warnings = load_manual_watchlist()
    print_manual_watchlist(rows, source, warnings)
    entered = input(
        "Select watchlist number, or type any ticker for manual entry: "
    ).strip()
    if not entered:
        entered = input("Manual ticker: ").strip()
    if entered.isdigit():
        index = int(entered)
        if 1 <= index <= len(rows):
            return rows[index - 1]["ticker"]
        raise ManualControlError("Watchlist selection is out of range.")
    symbol = entered.strip().upper()
    if not symbol:
        raise ManualControlError("Ticker is required.")
    return symbol


def print_sell_position_choices(positions: list[dict]) -> None:
    print("\nOWNED POSITIONS AVAILABLE TO SELL")
    if not positions:
        print("No positive long positions are currently available to select.")
        print("You may still type a ticker manually as a fallback.")
        return
    print(
        "#   SYMBOL     QTY OWNED    AVG COST    MARKET VALUE    "
        "UNREALIZED P&L    CUR  PRIMARY"
    )
    for index, item in enumerate(positions, 1):
        print(
            f"{index:<3} {item['symbol']:<9} {item['quantity']:>10g} "
            f"{(item['average_cost'] or 0):>11,.2f} "
            f"{(item['market_value'] or 0):>15,.2f} "
            f"{(item['unrealized_pnl'] or 0):>17,.2f} "
            f"{(item.get('currency') or '-'):<4} "
            f"{(item.get('primary_exchange') or '-'):<12}"
        )


def _select_sell_contract_from_positions(ib: IB, input_fn=input):
    positions = [
        position
        for position in get_positions(ib)
        if _plain_number(position.get("quantity")) is not None
        and _plain_number(position.get("quantity")) > 0
    ]
    print_sell_position_choices(positions)
    entered = input_fn(
        "Select owned position number, or type any ticker for manual entry: "
    ).strip()
    if not entered:
        entered = input_fn("Manual ticker: ").strip()
    if entered.isdigit():
        index = int(entered)
        if 1 <= index <= len(positions):
            return qualify_manual_contract(ib, positions[index - 1]["contract"])
        raise ManualControlError("Owned-position selection is out of range.")

    symbol = entered.strip().upper()
    if not symbol:
        raise ManualControlError("Ticker is required.")
    currency = input_fn("Currency [USD]: ").strip().upper() or "USD"
    return qualify_stock(ib, symbol, currency)


def _read_limit_price_with_default(prompt: str, suggested: float | None) -> float:
    if suggested is None:
        return _read_positive_number(prompt)
    entered = input(f"{prompt} [{suggested:g}]: ").strip()
    if not entered:
        return suggested
    try:
        value = float(entered)
    except Exception as error:
        raise ManualControlError("Limit price must be numeric.") from error
    if not math.isfinite(value) or value <= 0:
        raise ManualControlError("Limit price must be positive.")
    return value


def _select(items, prompt: str):
    if not items:
        raise ManualControlError("Nothing is available to select.")
    try:
        index = int(input(prompt).strip())
    except Exception as error:
        raise ManualControlError("Enter a valid menu number.") from error
    if index < 1 or index > len(items):
        raise ManualControlError("Selection is out of range.")
    return items[index - 1]


def _outside_hours_acknowledgement(status: dict) -> bool:
    if status["liquid_open"] is True:
        return False
    print("\n*** WARNING: MARKET LIQUID HOURS ARE CLOSED OR UNKNOWN ***")
    if status.get("detail"):
        print(status["detail"])
    print("The order may be rejected or held until a later session.")
    if not confirm_yes_no("Acknowledge outside-hours risk?"):
        raise ManualControlError("Outside-hours acknowledgement failed.")
    return True


def _interactive_order(ib: IB, action: str, order_type: str) -> None:
    if action == "BUY":
        symbol = _select_buy_symbol_from_watchlist()
        currency = input("Currency [USD]: ").strip().upper() or "USD"
        current_contract = qualify_stock(ib, symbol, currency)
    else:
        current_contract = _select_sell_contract_from_positions(ib)
        symbol = getattr(current_contract, "symbol", "")
    print(f"Resolved contract: {contract_identity_text(current_contract)}")
    held_quantity = max(0.0, current_broker_position(ib, current_contract))
    contract = current_contract
    price_info = get_current_market_price(ib, contract)
    current_price = price_info.get("price")
    print(
        "Current market price: "
        + (
            f"{current_price:,.4f} ({price_info.get('source')})"
            if current_price is not None
            else f"unavailable ({price_info.get('source')})"
        )
    )

    suggested_limit = (
        suggest_buy_limit_price(price_info)
        if action == "BUY" and order_type == "LIMIT"
        else None
    )
    limit_price = None
    if order_type == "LIMIT":
        if action == "BUY":
            limit_price = _read_limit_price_with_default(
                "Limit price", suggested_limit
            )
            if current_price is not None and limit_price > current_price * 1.2:
                print(
                    "WARNING: BUY limit is more than 20% above current market "
                    "price. The console will not refuse this solely for price, "
                    "but IBKR may reject it under broker precautionary checks."
                )
        else:
            limit_price = _read_limit_price_with_default(
                "Limit price",
                current_price,
            )

    quantity, used_all = _read_quantity(
        "Quantity"
        + (" (number or all): " if action == "SELL" else ": "),
        allow_all=(action == "SELL"),
        all_quantity=held_quantity,
    )
    status = get_market_hours_status(ib, contract)
    print_market_status(status, symbol, contract)
    outside_ack = _outside_hours_acknowledgement(status)
    estimated_notional = (
        quantity * (limit_price if limit_price is not None else current_price)
        if (limit_price is not None or current_price is not None)
        else None
    )

    print("\nORDER REVIEW")
    print(f"Action          : {action}")
    print(f"Order type      : {order_type}")
    print(f"Contract        : {contract_identity_text(contract)}")
    print(
        "Current price   : "
        + (f"{current_price:,.4f} ({price_info.get('source')})" if current_price is not None else "unavailable")
    )
    print(f"Limit price     : {limit_price:g}" if limit_price is not None else "Limit price     : N/A")
    print(f"Quantity        : {quantity:g}{' (all)' if used_all else ''}")
    print(
        "Est. notional   : "
        + (f"{estimated_notional:,.2f} USD" if estimated_notional is not None else "unavailable")
    )
    print(
        "Liquid hours    : "
        + ("OPEN" if status["liquid_open"] is True else "CLOSED" if status["liquid_open"] is False else "UNKNOWN")
    )
    if order_type == "MARKET":
        print("MARKET ORDER WARNING: fill price is not guaranteed.")
    if not confirm_yes_no("Confirm this order?"):
        raise ManualControlError("Order confirmation failed.")

    trade, outcome = place_manual_order(
        ib,
        contract,
        action,
        quantity,
        order_type,
        limit_price,
        acknowledge_outside_liquid_hours=outside_ack,
    )
    print(
        f"Order outcome: {outcome['category']}. "
        f"Source={outcome.get('outcome_source', 'unknown')} "
        f"ID={getattr(trade.order, 'orderId', None)} "
        f"PermID={getattr(trade.order, 'permId', None)} "
        f"IBKRStatus={outcome['ibkr_status']}"
    )
    if outcome.get("error_messages"):
        print("Broker messages:")
        for message in outcome["error_messages"]:
            print(f"- {message}")
    if outcome.get("audit_ok") is False:
        print("BROKER ACTION COMPLETED BUT AUDIT/LOCAL LOGGING FAILED")


MENU_OPTIONS = {
    "1": "Account Summary",
    "2": "Positions",
    "3": "Open Orders",
    "4": "BUY Limit",
    "5": "SELL Limit",
    "6": "BUY Market",
    "7": "SELL Market",
    "8": "Cancel Selected Order",
    "9": "Cancel All Orders",
    "10": "Liquidate Selected Position",
    "11": "Emergency Liquidate All Positions",
    "12": "Market/Liquid-Hours Status",
    "13": "Investable Capital Control",
    "14": "Latest Broker Execution History",
    "0": "Exit",
}


def print_menu() -> None:
    print("\n" + "=" * 72)
    print("MANUAL CONTROL CONSOLE")
    print("=" * 72)
    for key, label in MENU_OPTIONS.items():
        print(f"{key:>2}. {label}")


def execute_menu_choice(ib: IB, choice: str) -> bool:
    log_manual_action("MENU_SELECTION", choice=choice, label=MENU_OPTIONS.get(choice))

    if choice == "0":
        return False
    if choice == "1":
        print_account_summary(ib)
    elif choice == "2":
        print_positions(get_positions(ib))
    elif choice == "3":
        print_open_orders(get_open_trades(ib))
    elif choice == "4":
        _interactive_order(ib, "BUY", "LIMIT")
    elif choice == "5":
        _interactive_order(ib, "SELL", "LIMIT")
    elif choice == "6":
        _interactive_order(ib, "BUY", "MARKET")
    elif choice == "7":
        _interactive_order(ib, "SELL", "MARKET")
    elif choice == "8":
        trades = get_open_trades(ib)
        print_open_orders(trades)
        trade = _select(trades, "Order number to cancel: ")
        order_id = getattr(trade.order, "orderId", None)
        if not confirm_yes_no(f"Cancel order {order_id}?"):
            raise ManualControlError("Cancellation confirmation failed.")
        _, outcome = cancel_selected_order(ib, trade)
        print_order_outcome("Cancellation outcome", trade, outcome)
    elif choice == "9":
        print("WARNING: This cancels ALL API orders across client IDs.")
        if not confirm_yes_no("Cancel all API orders across client IDs?"):
            raise ManualControlError("Global cancellation confirmation failed.")
        _, outcome = cancel_all_orders(ib)
        print(
            f"Global cancellation outcome: {outcome['category']} | "
            f"remaining_open={outcome['remaining_open']}"
        )
    elif choice == "10":
        positions = get_positions(ib)
        print_positions(positions)
        position = _select(positions, "Position number to liquidate: ")
        status = get_market_hours_status(ib, position["contract"])
        print_market_status(status, position["symbol"], position["contract"])
        outside_ack = _outside_hours_acknowledgement(status)
        print(
            f"\nReview: MARKET liquidation for {position['symbol']} "
            f"quantity={abs(float(position['quantity'])):g} | "
            f"{contract_identity_text(position)}"
        )
        if not confirm_yes_no("Liquidate this selected position?"):
            raise ManualControlError("Liquidation confirmation failed.")
        trade, outcome = liquidate_position(
            ib,
            position,
            acknowledge_outside_liquid_hours=outside_ack,
        )
        print_order_outcome("Liquidation outcome", trade, outcome)
    elif choice == "11":
        positions = get_positions(ib)
        print_positions(positions)
        print("\nDANGER: This creates STOP_BOT.txt, globally cancels orders, and")
        print("submits MARKET liquidation orders for every current position.")
        if not confirm_yes_no("Emergency confirmation 1 of 2: liquidate all positions?"):
            raise ManualControlError("First emergency confirmation failed.")
        if not confirm_yes_no(
            "Emergency confirmation 2 of 2: I understand every position is targeted?"
        ):
            raise ManualControlError("Second emergency confirmation failed.")

        outside = False
        for position in positions:
            status = get_market_hours_status(ib, position["contract"])
            if status["liquid_open"] is not True:
                outside = True
                print_market_status(status, position["symbol"], position["contract"])
        outside_ack = _outside_hours_acknowledgement(
            {"liquid_open": False}
        ) if outside else False

        results = emergency_liquidate_all(
            ib,
            positions,
            acknowledge_outside_liquid_hours=outside_ack,
        )
        for result in results:
            if result["trade"] is not None:
                print_order_outcome(
                    f"{result['symbol']} emergency liquidation",
                    result["trade"],
                    result["outcome"],
                )
            else:
                print(
                    f"{result['symbol']}: FAILED | {result.get('error')} | "
                    f"{contract_identity_text(result)}"
                )
        failures = [item for item in results if not item["ok"]]
        if failures:
            print(
                "\nCRITICAL: one or more emergency liquidation requests failed "
                "or did not reach a safe broker-confirmed state."
            )
    elif choice == "12":
        mode, targets = _select_market_hours_targets(ib)
        show_market_hours_targets(
            ib,
            targets,
            continue_on_error=(mode == "ALL"),
        )
    elif choice == "13":
        run_investable_capital_control(ib)
    elif choice == "14":
        count = int(input("How many executions [20]: ").strip() or "20")
        rows, metadata = load_latest_execution_history(count)
        print_latest_execution_history(rows, metadata)
    else:
        print("Unknown menu option.")
    return True


def print_startup_warning() -> None:
    print("\n" + "!" * 78)
    print("PAPER TRADING MANUAL CONTROL CONSOLE")
    print("Verify IB Gateway is connected to the PAPER account before proceeding.")
    print("Orders here are sent directly and do not use automated strategy logic.")
    print("Manual trades do not update the strategy state file automatically.")
    print("!" * 78)


def run_console() -> int:
    print_startup_warning()
    ib = IB()
    try:
        connect_manual_console(ib)
        print(
            f"Connected to {cfg.HOST}:{cfg.PORT} with manual client ID "
            f"{MANUAL_CLIENT_ID}."
        )
        running = True
        while running:
            print_menu()
            choice = input("Select option: ").strip()
            try:
                running = execute_menu_choice(ib, choice)
            except BrokerActionError as error:
                log_manual_action(
                    "MENU_ACTION", "BROKER_FAILED", choice=choice, reason=str(error)
                )
                print(f"REJECTED BY IBKR OR BROKER/API FAILURE: {error}")
            except ManualControlError as error:
                log_manual_action(
                    "MENU_ACTION", "REFUSED", choice=choice, reason=str(error)
                )
                print(f"REFUSED BY CONSOLE PRE-CHECK: {error}")
            except Exception as error:
                log_manual_action(
                    "MENU_ACTION",
                    "FAILED",
                    choice=choice,
                    error_type=type(error).__name__,
                )
                print(f"ACTION FAILED: {type(error).__name__}: {error}")
        return 0
    except KeyboardInterrupt:
        log_manual_action("CONSOLE", "INTERRUPTED")
        print("\nConsole interrupted.")
        return 130
    except ManualControlError as error:
        print(f"Console startup failed: {error}")
        return 20
    finally:
        if ib.isConnected():
            ib.disconnect()
        log_manual_action("CONSOLE", "STOPPED")


if __name__ == "__main__":
    raise SystemExit(run_console())
